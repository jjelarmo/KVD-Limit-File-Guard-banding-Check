import os
import openpyxl

#   1. open all limit files
#   2. extract test name only. create nested lists: the first index is for the file, second index is for the testname
#   3. match files in 2 entries for common testnames. create a list for merged test names
#   4. find test elements and extract full electrical test information. create a dictionary: keys - testname, values - bin information in list
#   5. update dictionary with FT and QA limits
#   6. write updated dictionary in excel
#   7. compare limits and put flag on limist with no guardband

#   improvements - 7/19/2019
#   1. parse file names to check if FT and QA. arrange file names to be [FT, QA]
#   2. use test number as reference
#   3. check test_array to handle null list values
#   4. put labels on excel file

#   for final checkout
#   test criteria for working

def Guardband_Check():

    label1 = tk.Label(root, fg='green', font=('helvetica', 12, 'bold'))
    canvas1.create_window(150, 200, window=label1)
    
    path=os.getcwd()
    folder=os.fsencode(path)

    filename=[]
    for files in os.listdir(folder):
        files_decode=os.fsdecode(files)
        if files_decode.endswith('.lim'):
            filename.append(files_decode)

    for files in filename:
        for k in range(len(files)):
            if((files[k]== 'F') and (files[k+1]=='T')):
                FT=files
            if((files[k]== 'Q') and (files[k+1]=='A')):
                QA=files

    filename=[FT,QA]

    print(filename)
    
    testname_list=[]
    for files in filename:
        testname_list.append(testname_extractor(files))

    testname_merged=[]
    for test in testname_list[0]:
        if test in testname_list[1]:
            testname_merged.append(test)
    
    test_array=fulltest_extractor(filename[0])
    
    testname_full={}
    for test in testname_merged:
        for index in range(len(test_array)):
            if test_array[index][0] == test:
                temp_dict={test : test_array[index][0:3]}
                testname_full.update(temp_dict)

    del test_array
    
    test_array_FT=fulltest_extractor(filename[0])

    for test in testname_merged:
        for index in range(len(test_array_FT)):
            if test_array_FT[index][0] == test:
                testname_full[test].append(test_array_FT[index][4:6])

    del test_array_FT

    test_array_QA=fulltest_extractor(filename[1])

    for test in testname_merged:
        for index in range(len(test_array_QA)):
            if test_array_QA[index][0] == test:
                testname_full[test].append(test_array_QA[index][4:6])

    del test_array_QA
    del testname_merged
    del testname_list

    print('Write in excel')
    wb=openpyxl.Workbook()
    ws=wb.active

    ws.cell(row=1, column=4).value = 'FT lowerlimit'
    ws.cell(row=1, column=5).value = 'FT upperlimit'
    ws.cell(row=1, column=6).value = 'QA lowerlimit'
    ws.cell(row=1, column=7).value = 'QA upperlimit'
    ws.cell(row=1, column=8).value = 'Guardband Check Flag'
    
    index=2
    
    for keys in testname_full.keys():
        ws.cell(row=index, column=1).value = testname_full[keys][0]
        ws.cell(row=index, column=2).value = testname_full[keys][1]
        ws.cell(row=index, column=3).value = testname_full[keys][2]
        ws.cell(row=index, column=4).value = testname_full[keys][3][0]
        ws.cell(row=index, column=5).value = testname_full[keys][3][1]
        ws.cell(row=index, column=6).value = testname_full[keys][4][0]
        ws.cell(row=index, column=7).value = testname_full[keys][4][1]
        
        index=index+1

    length_index=index
    
    for index in range(2, length_index):
        print(index)
        if(isfloat(ws.cell(row=index, column=4).value)):
            FT_lowerlimit=float(ws.cell(row=index, column=4).value)
        elif(isint(ws.cell(row=index, column=4).value)):
            FT_lowerlimit=int(ws.cell(row=index, column=4).value)

        if(isfloat(ws.cell(row=index, column=5).value)):
            FT_upperlimit=float(ws.cell(row=index, column=5).value)
        elif(isint(ws.cell(row=index, column=5).value)):
            FT_lowerlimit=int(ws.cell(row=index, column=5).value)

        if(isfloat(ws.cell(row=index, column=6).value)):
            QA_lowerlimit=float(ws.cell(row=index, column=6).value)
        elif(isint(ws.cell(row=index, column=6).value)):
            QA_lowerlimit=int(ws.cell(row=index, column=6).value)

        if(isfloat(ws.cell(row=index, column=7).value)):
            QA_upperlimit=float(ws.cell(row=index, column=7).value)
        elif(isint(ws.cell(row=index, column=7).value)):
            QA_upperlimit=int(ws.cell(row=index, column=7).value)

        if((FT_lowerlimit>=QA_lowerlimit) and (FT_upperlimit<=QA_upperlimit) and ((QA_upperlimit - QA_lowerlimit)>=(FT_upperlimit-FT_lowerlimit)) and (FT_upperlimit>=FT_lowerlimit) and (QA_upperlimit>=QA_lowerlimit)):
           flag=0
        else:
           flag=1

        ws.cell(row=index, column=8).value=flag

    print('Finish parsing. Please check Guardband_Check excel file')   
    wb.save('Guardband_Check.xlsx')

    
def testname_extractor(files):
    #input: limit file
    #output: test name in list
    limitfile=open(files, "r")
    testlists=list(limitfile.readlines())
    
    FT_tests=[]
    
    for testlimits in testlists:
        if testlimits =='\n':
            continue
        else:
            testlimits_parsed=testlimits.split(',')
            FT_tests.append(testlimits_parsed[0])
        
    print('testname_extractor')
    return FT_tests

def fulltest_extractor(files):
    #input: limit file
    #output: array of electrical test information [testnumber, testname, unit, binning]
    limitfile=open(files, "r")
    testlists=list(limitfile.readlines())

    FT_tests=[]

    for testlimits in testlists:
        if testlimits =='\n':
            continue
        else:
            testlimits_parsed=testlimits.split(',')

            testnumber=testlimits_parsed[0]
            testname=testlimits_parsed[1]
            testunit=testlimits_parsed[2]
            testbin=testlimits_parsed[3]
            lowerlimit=testlimits_parsed[4]
            upperlimit=testlimits_parsed[5]

            FT_tests.append([testnumber,testname,testunit,testbin, lowerlimit, upperlimit])

    print('fulltest_extractor')
    return FT_tests

def isfloat(value):
    try:
        float(value)
        return True
    except ValueError:
        return False

def isint(value):
    try:
        int(value)
        return True
    except ValueError:
        return False
        
import tkinter as tk

root= tk.Tk()

canvas1 = tk.Canvas(root, width = 300, height = 300)
canvas1.pack()

instructions="""1. Create a local folder. Store this Guardband_Check.exe on your folder. \n
2. Store your limits file in your folder as well. Both limits file should have either FT or QA (in capital letters) on its file name.\n
3. Click Start_Check.\n
4. Wait for the program to finish parsing. Final output is a Guardband_Check excel file. It contains all common tests for FT and QA and Guarband Check flag for possible limit issues.\n"""
    
button1= tk.Button(text='Start Check',command=Guardband_Check, bg='brown',fg='white')
canvas1.create_window(150,150, window=button1)
T=tk.Text(root, height=10, width=200)
T.pack()
T.insert(tk.END, instructions)


root.mainloop()   
